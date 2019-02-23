
'''
import urllib.request
import urllib.parse
import json
'''



#(代码1)
'''
response = urllib.request.urlopen("http://www.fishc.com")
html = response.read()
html = html.decode("utf - 8")
print(html)
'''
#cross



#(代码2)
'''
response = urllib.request.urlopen("http://placekitten.com/g/200/300")
cat_img = response.read()
with open(cat_200_300.jpg,'wb') as f:
    f.write(cat_img)
'''
#error






#(代码3)
#初源代码
'''
url = "http://fanyi.youdao.com/translate?smartresult = dict&smartresult = rule&smartresult = ugc&sessionFrom = http://www.youdao.com/"
data = {}
data['type'] = 'AUTO'
data['i'] = 'I love fish'
data['doctype'] = 'json'
data['xmlVersion'] = '1.6'
data['keyfrom'] = 'fanyi.web'
data['ue'] = 'UTF - 8'
data['typoResult'] = 'true'
data = urllib.parse.urlencode(data).encode('utf - 8')
response = urllib.request.urlopen(url,data)
html = response.read().decode('utf - 8')
print(html)
'''




#(代码4---代码3改进)
'''
content = input("请输入需要翻译的内容：")
url = "http://fanyi.youdao.com/translate?smartresult = dict&smartresult = rule&smartresult = ugc&sessionFrom = http://www.youdao.com/"
data = {}
data['type'] = 'AUTO'
data['i'] = content
data['doctype'] = 'json'
data['xmlVersion'] = '1.6'
data['keyfrom'] = 'fanyi.web'
data['ue'] = 'UTF - 8'
data['typoResult'] = 'true'
data = urllib.parse.urlencode(data).encode('utf - 8')
response = urllib.request.urlopen(url,data)
html = response.read().decode('utf - 8')
target = json.loads(html)
print("翻译结果：%s" % (target['translateResult'][0][0]['tgt']))
'''
#不合法被截断





#(代码5---代码3改进)
'''
content = input("请输入需要翻译的内容：")
url = "http://fanyi.youdao.com/translate?smartresult = dict&smartresult = rule&smartresult = ugc&sessionFrom = http://www.youdao.com/"
#url = "http://fanyi.youdao.com/translate_o?smartresult=dict&smartresult=rule"
#head = {}
#head['Referer'] = 'http://fanyi.youdao.com'
#head['User - Agent'] = 'Mozilla/5.0(Macintosh;Intel Mac OS X 10_10_1) AppleWebKit/537.36(KHTML,like Gecko)Chrome/39.0.2171.95 Safari/537.36X - Requested - With:XMLHttpRequest'
#head['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36'
data = {}
data['type'] = 'AUTO'
data['i'] = content
data['doctype'] = 'json'
data['xmlVersion'] = '1.6'
data['keyfrom'] = 'fanyi.web'
data['ue'] = 'UTF - 8'
data['typoResult'] = 'true'
data = urllib.parse.urlencode(data).encode('utf - 8')
#req = urllib.request.Request(url,data,head)

req = urllib.request.Request(url,data)
req.add_header('Referer','http://fanyi.youdao.com')
req.add_header('User-Agent','Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36')

response = urllib.request.urlopen(req)
html = response.read().decode('utf-8')
target = json.loads(html)
print("翻译结果：%s" % (target['translateResult'][0][0]['tgt']))
'''
#不合法被截断











#(代码6---猫眼电影的爬虫---结果保存在一个文本中---运行成功)
'''
import json
import requests
from requests.exceptions import RequestException
import re
import time

def get_one_page(url):
    try:
        headers = {'User - Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36'}
        response = requests.get(url,headers=headers)
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        return None

def parse_one_page(html):
    pattern = re.compile('<dd>.*?board-index.*?>(\d+)</i>.*?data-src="(.*?)".*?name"><a'
                         +'.*?>(.*?)</a>.*?star">(.*?)</p>.*?releasetime">(.*?)</p>'
                         +'.*?integer">(.*?)</i>.*?fraction">(.*?)</i>.*?</dd>',re.S)
    items = re.findall(pattern,html)
    for item in items:
        yield{
            'index':item[0],
            'image':item[1],
            'title':item[2],
            'actor':item[3].strip()[3:],
            'time':item[4].strip()[5:],
            'score':item[5]+item[6]
        }

def write_to_file(content):
    with open('result.txt','a',encoding='utf-8') as f:
        f.write(json.dumps(content,ensure_ascii=False)+'\n')

def main(offset):
    url = 'http://maoyan.com/board/4?offset='+str(offset)
    html = get_one_page(url)
    for item in parse_one_page(html):
        print(item)
        write_to_file(item)

if __name__ == '__main__':
    for i in range(10):
        main(offset=i*10)
        time.sleep(1)

'''
















#(代码7---一个爬虫---结果有bug未改成功)
'''
import json
import os
from urllib.parse import urlencode
import pymongo
import requests
from bs4 import BeautifulSoup
from requests.exceptions import ConnectionError
import re
from multiprocessing import Pool
from hashlib import md5
from json.decoder import JSONDecodeError
from config import *

client = pymongo.MongoClient(MONGO_URL, connect=False)
db = client[MONGO_DB]


def get_page_index(offset, keyword):
    data = {
        'autoload': 'true',
        'count': 20,
        'cur_tab': 3,
        'format': 'json',
        'keyword': keyword,
        'offset': offset,
    }
    params = urlencode(data)
    base = 'http://www.toutiao.com/search_content/'
    url = base + '?' + params
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None


def download_image(url):
    print('Downloading', url)
    try:
        response = requests.get(url)
        if response.status_code == 200:
            save_image(response.content)
        return None
    except ConnectionError:
        return None


def save_image(content):
    file_path = '{0}/{1}.{2}'.format(os.getcwd(), md5(content).hexdigest(), 'jpg')
    print(file_path)
    if not os.path.exists(file_path):
        with open(file_path, 'wb') as f:
            f.write(content)
            f.close()


def parse_page_index(text):
    try:
        data = json.loads(text)
        if data and 'data' in data.keys():
            for item in data.get('data'):
                yield item.get('article_url')
    except JSONDecodeError:
        pass


def get_page_detail(url):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None


def parse_page_detail(html, url):
    soup = BeautifulSoup(html, 'lxml')
    result = soup.select('title')
    title = result[0].get_text() if result else ''
    images_pattern = re.compile('gallery: JSON.parse\("(.*)"\)', re.S)
    result = re.search(images_pattern, html)
    if result:
        data = json.loads(result.group(1).replace('\\', ''))
        if data and 'sub_images' in data.keys():
            sub_images = data.get('sub_images')
            images = [item.get('url') for item in sub_images]
            for image in images: download_image(image)
            return {
                'title': title,
                'url': url,
                'images': images
            }


def save_to_mongo(result):
    if db[MONGO_TABLE].insert(result):
        print('Successfully Saved to Mongo', result)
        return True
    return False


def main(offset):
    text = get_page_index(offset, KEYWORD)
    urls = parse_page_index(text)
    for url in urls:
        html = get_page_detail(url)
        result = parse_page_detail(html, url)
        if result: save_to_mongo(result)


if __name__ == '__main__':
    pool = Pool()
    groups = ([x * 20 for x in range(GROUP_START, GROUP_END + 1)])
    pool.map(main, groups)
    pool.close()
    pool.join()


'''














#(代码8---一个爬虫---结果有bug未改成功)
'''
import requests
import re
import os
import sys


def get_downloadurl():
    search_url = 'http://yyets.mirrors.ga/search/index?keyword=%s' % sys.argv[1]
    resource_url = 'http://yyets.mirrors.ga/resource/'
    r = requests.get(search_url, timeout=20)
    data = r.content
    match = re.search(r'yyets\.mirrors\.ga/resource/(\d+)', data)
    if match:
        num = match.group(1)
        resource_url = resource_url + num
        r = requests.get(resource_url, timeout=20)
        data = r.content
        download_list = re.findall(r'"(ed2k://\S+.(S\d+E\d+)\S+1024\w576\S+)"', data)
        if download_list:
            with open('result.txt', 'w') as f:
                for i in download_list:
                    while download_list.count(i) > 1:
                        del download_list[download_list.index(i)]
                    f.write('%s %s\n%s\n' % (sys.argv[1], i[1], i[0]))
                print ('Save as result.txt')
        else:
            print('No Resource')
    else:
        print ('Not Found')


def main():
    filename = sys.argv[0]
    if len(sys.argv) < 2:
        print( 'Usage: ' + os.path.basename(filename) + ' keyword')
        sys.exit()
    get_downloadurl()

if __name__ == '__main__':
    main()
'''



#(代码9---爬虫---此爬虫代码对计算要求较高---未运行---观察其中结构)
import sys
import time
import urllib
#import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
import importlib

#reload(sys)
importlib.reload(sys)
#sys.setdefaultencoding('utf8')

# Some User Agents
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}, \
       {
           'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'}, \
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num = 0;
    book_list = []
    try_times = 0

    while (1):
        # url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        url = 'http://www.douban.com/tag/' + urllib.request.quote(book_tag) + '/book?start=' + str(page_num * 15)
        time.sleep(np.random.rand() * 5)

        # Last Version
        try:
            req = urllib.request(url, headers=hds[page_num % len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text = str(source_code)
       # except (urllib2.HTTPError, urllib2.URLError), e:
        except urllib.error.URLError as e:
            print
            e
            continue

        ##Previous Version, IP is easy to be Forbidden
        # source_code = requests.get(url)
        # plain_text = source_code.text

        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('div', {'class': 'mod book-list'})

        try_times += 1;
        if list_soup == None and try_times < 200:
            continue
        elif list_soup == None or len(list_soup) <= 1:
            break  # Break when no informatoin got after 200 times requesting

        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class': 'title'}).string.strip()
            desc = book_info.find('div', {'class': 'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class': 'title'}).get('href')

            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).string.strip()
            except:
                rating = '0.0'
            try:
                # people_num = book_info.findAll('span')[2].string.strip()
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num = '0'

            book_list.append([title, rating, people_num, author_info, pub_info])
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print
        'Downloading Information From Page %d' % page_num
    return book_list


def get_people_num(url):
    # url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib2.Request(url, headers=hds[np.random.randint(0, len(hds))])
        source_code = urllib2.urlopen(req).read()
        plain_text = str(source_code)
    #except (urllib2.HTTPError, urllib2.URLError), e:
    except :
        pass
    soup = BeautifulSoup(plain_text)
    people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))  # utf8->unicode
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1
    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i].decode())
    save_path += '.xlsx'
    wb.save(save_path)


if __name__ == '__main__':
    # book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    # book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    # book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    # book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    # book_tag_lists = ['数学']
    # book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    # book_tag_lists = ['商业','理财','管理']
    # book_tag_lists = ['名著']
    # book_tag_lists = ['科普','经典','生活','心灵','文学']
    # book_tag_lists = ['科幻','思维','金融']
    book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists, book_tag_lists)